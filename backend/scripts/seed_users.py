"""Seeds a large batch of demo user accounts across every LLM-RBAC role (see
docs/ROLE_PERMISSION_MATRIX.md) directly into Postgres — bypassing POST
/users, which always creates role="user" (there's no self-service way to
become HR/Project Manager/Admin), and bypassing document ingestion entirely
(out of scope for this script by design — see the conversation this was
requested in: user accounts only, no documents).

Emails are short and predictable — empN/hrN/pmN/ceoN@mail.com — rather than
realistic-looking names, so they're easy to type/remember when testing role-
based UI by hand. Every seeded account shares one password (DEMO_PASSWORD)
for the same reason: this is throwaway dev/demo data, never a production
credential, and sharing one password+hash across every row is a deliberate
shortcut for that reason (never do this for real accounts).

Usage (from backend/):
    python -m scripts.seed_users
    python -m scripts.seed_users --employees 40 --hr 10 --project-managers 10 --admins 3

Idempotent per role — re-running tops up to the requested counts, numbering
new accounts on from whatever's already seeded, without duplicating anyone.
Writes scripts/seeded_users.csv and scripts/seeded_users.xlsx (email, role,
display name, password) as login-credential references; both are gitignored,
never commit them.
"""

import argparse
import csv
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from sqlalchemy.orm import Session  # noqa: E402

from app.core.roles import Role  # noqa: E402
from app.db.postgres import ensure_schema, new_session  # noqa: E402
from app.models.user import UserModel  # noqa: E402
from app.services.auth.password import hash_password  # noqa: E402
from app.services.llm_rbac.policy_loader import role_config  # noqa: E402

DEMO_PASSWORD = "Demo@12345"
EMAIL_DOMAIN = "mail.com"
OUTPUT_CSV = Path(__file__).resolve().parent / "seeded_users.csv"
OUTPUT_XLSX = Path(__file__).resolve().parent / "seeded_users.xlsx"

# Short login-friendly prefix per role. CEO was split out from Admin
# (previously one combined "CEO/Admin" role, hence admin's old "ceo" prefix)
# — each now gets its own, matching its own Role enum value.
ROLE_PREFIXES = {
    Role.USER.value: "emp",
    Role.HR.value: "hr",
    Role.PROJECT_MANAGER.value: "pm",
    Role.CEO.value: "ceo",
    Role.ADMIN.value: "admin",
}


def _next_index(db: Session, prefix: str) -> int:
    """Lowest unused N for f"{prefix}{N}@{EMAIL_DOMAIN}" — scans existing
    seeded emails for this prefix rather than trusting a per-role row count,
    so a run that only adds e.g. --hr doesn't collide if earlier runs left
    gaps (a manually deleted account, a differently-counted prior run)."""
    pattern = re.compile(rf"^{re.escape(prefix)}(\d+)@{re.escape(EMAIL_DOMAIN)}$")
    existing = db.query(UserModel.email).filter(UserModel.email.like(f"{prefix}%@{EMAIL_DOMAIN}")).all()
    max_index = 0
    for (email,) in existing:
        match = pattern.match(email)
        if match:
            max_index = max(max_index, int(match.group(1)))
    return max_index + 1


def seed(counts: dict[str, int]) -> list[tuple[str, str, str]]:
    """Returns every (email, role, display_name) currently seeded for these
    roles — not just what this run inserted — so the CSV reference always
    reflects the database, not one run's delta."""
    ensure_schema()
    db = new_session()
    # One hash for every seeded row (see module docstring) — the expensive
    # part of hash_password is the 390k-iteration PBKDF2 call; computing it
    # per-row would turn a 100-account seed into a multi-minute run for no
    # benefit on data nobody should treat as a real credential.
    password_hash = hash_password(DEMO_PASSWORD)

    created = 0
    try:
        for role, count in counts.items():
            if count <= 0:
                continue
            prefix = ROLE_PREFIXES[role]
            department = role_config(role).department_default
            start = _next_index(db, prefix)
            for i in range(start, start + count):
                db.add(UserModel(
                    email=f"{prefix}{i}@{EMAIL_DOMAIN}", display_name=f"{role_config(role).display_name} {i}",
                    password_hash=password_hash, role=role, department=department, is_active=True,
                ))
                created += 1
        db.commit()

        # Match only this script's own empN/hrN/pmN/ceoN@{domain} naming —
        # NOT every account on this domain. A plain "%@{domain}" filter would
        # also sweep up anyone who happens to have manually created an
        # account like admin@mail.com or hr@mail.com themselves, and falsely
        # claim (in the CSV below) that DEMO_PASSWORD is their password.
        name_pattern = "|".join(re.escape(p) for p in ROLE_PREFIXES.values())
        all_seeded = (
            db.query(UserModel)
            .filter(UserModel.email.op("~")(rf"^({name_pattern})[0-9]+@{re.escape(EMAIL_DOMAIN)}$"))
            .order_by(UserModel.role, UserModel.display_name)
            .all()
        )
        rows = [(u.email, u.role, u.display_name or "") for u in all_seeded]
    finally:
        db.close()

    print(f"Created {created} new user(s) this run.")
    return rows


def _write_xlsx(rows: list[tuple[str, str, str]], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Credentials"

    headers = ["Email", "Role", "Display Name", "Password"]
    ws.append(headers)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")

    for email, role, display_name in rows:
        ws.append([email, role_config(role).display_name, display_name, DEMO_PASSWORD])

    for col, width in enumerate([30, 16, 20, 16], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--employees", type=int, default=70, help="role=user (Employee) accounts to create")
    parser.add_argument("--hr", type=int, default=15, help="role=hr accounts to create")
    parser.add_argument("--project-managers", type=int, default=15, help="role=project_manager accounts to create")
    parser.add_argument("--ceos", type=int, default=5, help="role=ceo accounts to create")
    parser.add_argument("--admins", type=int, default=2, help="role=admin accounts to create")
    args = parser.parse_args()

    rows = seed({
        Role.USER.value: args.employees,
        Role.HR.value: args.hr,
        Role.PROJECT_MANAGER.value: args.project_managers,
        Role.CEO.value: args.ceos,
        Role.ADMIN.value: args.admins,
    })

    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["email", "role", "display_name", "password"])
        for email, role, display_name in rows:
            writer.writerow([email, role, display_name, DEMO_PASSWORD])
    _write_xlsx(rows, OUTPUT_XLSX)

    by_role: dict[str, list[str]] = {}
    for email, role, _ in rows:
        by_role.setdefault(role, []).append(email)

    print(f"\n{len(rows)} total @{EMAIL_DOMAIN} account(s) in the database:")
    for role, emails in by_role.items():
        print(f"  {role_config(role).display_name:<16} {len(emails):>4}   e.g. {emails[0]}")
    print(f"\nShared password for every seeded account: {DEMO_PASSWORD}")
    print(f"Full list written to {OUTPUT_CSV}")
    print(f"Full list written to {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
